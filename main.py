import xlwings
import tkinter as tk
from tkinter import filedialog
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.borders import Border, Side
import os
from copy import copy
import datetime


def save_table(workbook):
    f = filedialog.asksaveasfile(
    	title='Выберите исходный файл', 
    	defaultextension=".xlsx", 
    	filetypes=[('Файлы Excel', '*.xlsx'), ('Все файлы', '*.*')])
    if f is None:
    	if tk.messagebox.askyesno('Файл не выбран!', 'Повторить?'):
        	save_table(workbook)
    	else:
    		tk.messagebox.showwarning('Не сохранено', 'Файл не был сохранён.')
    	return
    filename = f.name
    f.close()
    workbook.save(filename)

def open_workbook(file_path=None, data_only=False):
	if file_path is None:
		file_path = filedialog.askopenfilename(
			title='Выберите исходный файл', 
	    	defaultextension=".xlsx", 
	    	filetypes=[('Файлы Excel', '*.xlsx'), ('Все файлы', '*.*')]
    	)

	if (file_path == '' or os.path.isfile(file_path) == False):
		return None

	return load_workbook(file_path, data_only=data_only)

def read_column(index, sheet):
	if (index >= sheet.max_column):
		return None
	return [[cell.value if cell.value is not None else 0 for cell in col] for col in sheet.columns][index]

def read_source(source, sheetnames):
	sheet = source[sheetnames[0]]
	title_column = read_column(0, sheet)[:17:]
	if (title_column != ['Вещество', 'Азот (IV) оксид (Азота диоксид)', 'Азот (II) оксид (Азота оксид)', 'Углерод (Сажа)', 'Сера диоксид (Ангидрид сернистый)', 'Дигидросульфид (Сероводород)', 'Углерод оксид', 'Углерод диоксид', 'Метан', 'Этан', 'Пропан', 'Бутан', 'Пентан', 'Смесь углеводородов предельных С6-С10', 'Бенз/а/пирен (3,4-Бензпирен)', 'Метантиол (Метилмеркаптан)', 'Этантиол (Этилмеркаптан)']):
		tk.messagebox.showerror(
			'Ошибка струтуры файла', 
			'Сверьтесь с шаблоном!\nНе верные порядок или названия веществ!\n' + 
			f"Лист: '{sheet.title}'."
		)
		return None

	second_column = read_column(1, sheet)[:17:]
	if (second_column != ['Мера', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год', 'тонн/год']):
		tk.messagebox.showerror(
			'Ошибка струтуры файла', 
			'Сверьтесь с шаблоном!\nНе верная Мера!\n' + 
			f"Лист: '{sheet.title}'."
		)
		return None

	materials = {}
	index = 2
	while (column := read_column(index, sheet)):
		machine = column[0]
		if (len(column) < 17):
			tk.messagebox.showerror(
				'Ошибка струтуры файла', 
				'Сверьтесь с шаблоном!\n' +
				f"Для '{machine}' указано недостаточное число строк!\n" +
				f"Лист: '{sheet.title}'."
			)
			return None

		if (column[0] in materials):
			tk.messagebox.showerror(
				'Ошибка имён', 
				'Добавте уникальности!\n' +
				f"Установка '{machine}' уже встречалось ранее!\n" +
				f"Используйте например {machine}(1).\n"+
				f"Лист: '{sheet.title}'."
			)
			return None
		materials.update({machine: column[1:17:]})
		index+=1

	sheet = source[sheetnames[1]]
	years = read_column(0, sheet)[2::]
	while (years[-1] == 0):
		years.pop()

	if len(years) < 3:
		tk.messagebox.showerror(
			'Недостаточно данных',
			f"Необходимо минимум 3 основных года!\n" +
			f"Лист: '{sheet.title}'."
		)
		return None

	min_year = min(years[3:])
	if years[0] >= min_year:
		tk.messagebox.showerror(
				'Года идут не по порядку', 
				f"{years[0]}г. должен был быть минимальным!\n"+
				f"Конфликтный год: {min_year}\n"+
				f"Лист: '{sheet.title}'."
			)
		return None

	if years[1] != years[2]:
		tk.messagebox.showerror(
				'Года идут не по порядку', 
				f"Строки 4 и 5 должны содержать одинаковые года (до выработки и после)\n" +
				f"Лист: '{sheet.title}'."
			)
		return None

	if years[1] >= min_year:
		tk.messagebox.showerror(
				'Года идут не по порядку', 
				f"{years[1]}г. должен был быть меньше последующих!\n"+
				f"Конфликтный год: {min_year}\n"+
				f"Лист: '{sheet.title}'."
			)
		return None

	start = years[3]
	for i in range(3, len(years)):
		if (years[i] == 0):
			tk.messagebox.showerror(
				'Год не указан', 
				"Вы не указали год!\n" + 
				f"Строка: {i+1}\n"+
				f"Лист: '{sheet.title}'."
			)
			return None

		if (years[i] != start + i):
			tk.messagebox.showwarning(
				'Года идут не по порядку', 
				f"Ожидался {start + i}г. а был получен {years[i]}г.\n"+
				f"Строка: {i+1}\n"+
				f"Лист: '{sheet.title}'."
			)
			start = years[i] - i

	sorted_years = years
	sorted_years.sort()
	if (years != sorted_years):
		tk.messagebox.showwarning(
			'Года идут не по позрастанию', 
			f"Лист: '{sheet.title}'."
		)

	result = {}
	index = 1
	voluem_exist = True
	while (prodaction := read_column(index, sheet)) and (burning := read_column(index+1, sheet)):
		machine = prodaction[0]
		if machine in result:
			tk.messagebox.showerror(
				'Ошибка имён', 
				'Добавте уникальности!\n' +
				f"Установка '{machine}' уже встречалось ранее!\n" +
				f"Используйте например '{machine}(1)'.\n"+
				f"Лист: '{sheet.title}'."
			)
			return None

		if machine not in materials:
			tk.messagebox.showerror(
				'Ошибка имён', 
				f"Вещества для '{machine}' не были указаны!\n" +
				f"Лист: '{sheetnames[0]}'."
			)
			voluem_exist = False

		result.update({machine: {
			'materials': materials[machine] if voluem_exist else None,
			'prodaction': prodaction[2:len(years)+2:], 
			'burning': burning[2:len(years)+2:]
		}})
		index+=2

	materials_exist = True
	for key in materials:
		if key not in result:
			tk.messagebox.showerror(
				'Ошибка имён', 
				f"Объёмы для установки '{key}' отсутствуют!\n" +
				f"Лист: '{sheet.title}'."
			)
			materials_exist = False

	if materials_exist and voluem_exist:
		return result, years
	return None

def write_year(sheet, column, year, voluems):
	curr_col = get_column_letter(column)

	sheet[f"{curr_col}1"].value = year
	sheet[f"{curr_col}2"].value = voluems[0] if voluems is not None else 0
	sheet[f"{curr_col}3"].value = voluems[1] if voluems is not None else 0

	# extend tamplate
	if column > 5:
		before_col = get_column_letter(column-1)
		sheet.column_dimensions[curr_col].width = 17
		for row in range(1, 5):
			sheet[f"{curr_col}{row}"]._style = copy(sheet[f"{before_col}{row}"]._style)
		sheet[f"{curr_col}4"] = Translator(sheet[f"{before_col}4"].value, f"{before_col}4").translate_formula(f"{curr_col}4")

def extend_sheet(sheet, start_col, heigth, count, col_span=1, translate_dist = 1, col_width=15, start_row=1):
	# Remove tamplate column
	if (count <= 0):
		row_index = 1
		curr_col = get_column_letter(start_col)
		end_col = get_column_letter(start_col + col_span) 
		for row in sheet[f"{curr_col}{start_row}:{end_col}{heigth}"]:
			end_cel_name = f"{get_column_letter(start_col + col_span)}{row_index}"
			row[0]._style = copy(sheet[end_cel_name]._style)
			row[0].value = None
			row_index += 1
			for cell in row[1::]:
				outside_cell_name = f"{get_column_letter(start_col + col_span + 1)}1"
				cell._style = copy(sheet[outside_cell_name]._style)
				cell.value = None
		return

	# One (template) column exist, not need to copy it
	if (count == 1):
		return

	for iteration in range(1, count):
		begin_index = start_col + iteration*col_span 				# start of extending area
		begin_col = get_column_letter(begin_index)					# name of this column
		end_col = get_column_letter(begin_index + col_span -1)		# end of extending area
		area = f"{begin_col}{start_row}:{end_col}{heigth}"
		#print(f"Area is: {area}")
		row_index = start_row
		for row in sheet[area]:
			cell_index = 0
			for cell in row:
				curr_col = get_column_letter(begin_index + cell_index)
				next_col = get_column_letter(begin_index + cell_index + 1)
				distance = translate_dist if type(translate_dist) is int else translate_dist[row_index-1]

				translate_col = get_column_letter(start_col + (iteration + cell_index) * distance)
				sheet.column_dimensions[curr_col].width = col_width

				temp_cell_name = f"{get_column_letter(start_col + cell_index)}{row_index}"
				translate_cell_name = f"{translate_col}{row_index}"
				temp_cell = sheet[temp_cell_name]
				
				#print(f"Translate from {temp_cell_name} to {translate_cell_name}, and past in {curr_col}{row_index}")

				cell._style = copy(temp_cell._style)
				cell.border = copy(temp_cell.border)

				sheet[f"{curr_col}{row_index}"] = Translator (
					temp_cell.value, 				# from template
					temp_cell_name					# at cell
				).translate_formula (
					translate_cell_name 			# Translate to
				)
				cell_index += 1
			row_index += 1

def make_report(directory, col_count):
	row_count = 9
	sheetnames = ['Исходные данные', 'Объём сжигания', 'Лимит сжигания 5%', 'Сверх нормативное сжигание', 'Лимиты', 'Сверх лимит (вред)', 'Итоговый вывод']
	
	excel_app = xlwings.App(visible=False)

	report = Workbook()
	r_sheet = report.active
	r_sheet.title = 'Общая сводка'

	for dirpath, dirnames, filenames in os.walk(directory):
		counter = 0
		for filename in [f for f in filenames if f.endswith(".xlsx")]:
			path = os.path.join(dirpath, filename)
			print(path)
			excel_book = excel_app.books.open(path)
			excel_book.save()
			excel_book.close()
			temp_report = open_workbook(path, data_only=True)
			if (temp_report.sheetnames != sheetnames):
				tk.messagebox.showwarning(
					'Ошибка струтуры файла',
					'В промежуточном файле не верные порядок или названия листов!\n' +
					f"Файл: {path}"
				)
				temp_report.close()
				continue

			sheet = temp_report[sheetnames[-1]]
			for col_number in range(1, col_count + 1):
				column_name = get_column_letter(col_number)
				r_sheet.column_dimensions[column_name].width = 35 if col_number == 1 else 15
				for row_index in range(1, row_count + 1):
					r_sheet.row_dimensions[row_index + (row_count+1) * counter].height  = 30 if row_index == 1 else 15

					r_sheet[f"{column_name}{row_index + (row_count+1) * counter}"].value = copy(sheet[f"{column_name}{row_index}"].value)
					target_cell = r_sheet[f"{column_name}{row_index + (row_count+1) * counter}"]
					source_cell = sheet[f"{column_name}{row_index}"]

					target_cell.font = copy(source_cell.font)
					target_cell.border = copy(source_cell.border)
					target_cell.fill = copy(source_cell.fill)
					target_cell.number_format = copy(source_cell.number_format)
					target_cell.protection = copy(source_cell.protection)
					target_cell.alignment = copy(source_cell.alignment)

			temp_report.close()
			counter += 1

	save_table(report)
	excel_app.quit()


def main():
	root = tk.Tk()
	root.withdraw()
	source = open_workbook(sys.argv[1] if len(sys.argv) == 2 else None, data_only=True)
	if (source is None):
		tk.messagebox.showwarning('Файл не выбран', 'Вы не выбрали файл!')
		return

	sheetnames = ['Объём по веществам', 'Объём по годам']

	if source.sheetnames != sheetnames:
		tk.messagebox.showerror(
			'Ошибка струтуры файла', 
			'Сверьтесь с шаблоном!\nНе верные порядок или названия листов!'
		)
		return

	source_data = read_source(source, sheetnames)
	source.close()
	if (source_data is None):
		return

	values, years = source_data
	start_year, end_year = min(years[3::]), max(years)
	sheetnames = ['Исходные данные', 'Объём сжигания', 'Лимит сжигания 5%', 'Сверх нормативное сжигание', 'Лимиты', 'Сверх лимит (вред)', 'Итоговый вывод']

	now = datetime.datetime.now().strftime("%d_%m_%Y")
	temp_dir = f"{now}_temp"
	if not os.path.exists(temp_dir):
		os.makedirs(temp_dir)

	for machine in values:
		data = values[machine]
		template = open_workbook('template.xlsx')
		if (template is None):
			tk.messagebox.showerror('Шаблон не найден', "Не найден файл 'template.xlsx'!")
			return

		if template.sheetnames != sheetnames:
			tk.messagebox.showerror(
				'Ошибка струтуры файла', 
				'Сверьтесь с шаблоном!\nНе верные порядок или названия листов!'
			)
			return


		sheet = template[sheetnames[0]]
		sheet.cell(row=1, column=1).value = machine

		# first 3 years by hand
		write_year(sheet, 3, years[0], (
			data['prodaction'][0]*1000000, data['burning'][0]*1000000
		))
		write_year(sheet, 4, f"До выраб. {years[1]}", (
			data['prodaction'][1]*1000000, data['burning'][1]*1000000
		))
		write_year(sheet, 5, f"После выраб.{years[2]}", (
			data['prodaction'][2]*1000000, data['burning'][2]*1000000
		))

		column = 6
		extend_count = 0
		for year in range(start_year, end_year+1):
			extend_count += 1
			voluems = None
			if (year in years):
				index = years.index(year)
				voluems = data['prodaction'][index]*1000000, data['burning'][index]*1000000
			write_year(sheet, column, year, voluems)
			column += 1


		sheet = template[sheetnames[1]]
		row = 2
		for material_voluem in data['materials']:
			sheet[f"C{row}"].value = material_voluem
			row+=1

		extend_sheet(template[sheetnames[1]], 6, 18, extend_count)
		extend_sheet(template[sheetnames[2]], 6, 18, extend_count)
		extend_sheet(template[sheetnames[3]], 4, 18, extend_count)
		extend_sheet(template[sheetnames[4]], 7, 18, extend_count, 2) 				# Парная симетрия для формул
		extend_sheet(template[sheetnames[4]], 7, 19, extend_count*2, start_row=19) 	# Унарная симетри для сумм
		extend_sheet(template[sheetnames[5]], 4, 18, extend_count)
		template[sheetnames[6]]["A1"] = machine
		extend_sheet(template[sheetnames[6]], 6, 9, extend_count, 
			translate_dist=[1, 1, 2, 2, 1, 1, 1, 1, 1]
		)

		template.save(f'{temp_dir}/{machine}.xlsx')

	make_report(temp_dir, end_year - years[0] + 2)



if __name__ == '__main__':
	main()